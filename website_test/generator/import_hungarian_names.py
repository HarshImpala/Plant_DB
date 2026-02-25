"""
Import Hungarian common names from the tropical_test Excel file into plants.db.

Updates `plants.common_name_hungarian` by matching Latin names against
input/scientific/canonical plant names (exact first, then genus+species).
"""

import re
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent.parent
DB_PATH = BASE_DIR / "data" / "plants.db"
EXCEL_PATH = (
    BASE_DIR.parent
    / "old_scripts"
    / "excel_files"
    / "tropical_test"
    / "tropusi_haszon_test.xlsx"
)


def _norm(text: str | None) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[\u2018\u2019'\"`]", "", text)
    text = re.sub(r"\s+", " ", text)
    return text


def _binomial(text: str | None) -> str:
    """Reduce scientific name text to `genus species`."""
    text = (text or "").strip()
    text = re.sub(r"'[^']+'", "", text)  # cultivar names
    text = re.sub(r"\s*\([^)]*\)", "", text)  # parenthesized author/comments
    text = re.sub(r"\s+", " ", text).strip()
    parts = text.split(" ")
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1]}".lower()
    return text.lower()


def ensure_column(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(plants)")
    columns = {row[1] for row in cur.fetchall()}
    if "common_name_hungarian" not in columns:
        cur.execute("ALTER TABLE plants ADD COLUMN common_name_hungarian TEXT")
        conn.commit()


def read_excel_rows(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        latin = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        hungarian = (row[2] or "").strip() if len(row) > 2 and row[2] else ""
        if not latin or not hungarian:
            continue
        rows.append((latin, hungarian))
    return rows


def main() -> None:
    print(f"DB: {DB_PATH}")
    print(f"Excel: {EXCEL_PATH}")

    if not DB_PATH.exists():
        raise SystemExit(f"Database not found: {DB_PATH}")
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Excel file not found: {EXCEL_PATH}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    ensure_column(conn)
    cur = conn.cursor()

    cur.execute("SELECT id, input_name, scientific_name, canonical_name FROM plants")
    plants = cur.fetchall()

    exact_index: dict[str, set[int]] = {}
    binomial_index: dict[str, set[int]] = {}

    for plant in plants:
        pid = plant["id"]
        for candidate in (
            plant["input_name"],
            plant["scientific_name"],
            plant["canonical_name"],
        ):
            k_exact = _norm(candidate)
            if k_exact:
                exact_index.setdefault(k_exact, set()).add(pid)
            k_bin = _binomial(candidate)
            if k_bin:
                binomial_index.setdefault(k_bin, set()).add(pid)

    rows = read_excel_rows(EXCEL_PATH)
    print(f"Excel rows with Hungarian names: {len(rows)}")

    matched_unique = 0
    matched_multi = 0
    unmatched = 0
    updated = 0

    for latin_name, hungarian_name in rows:
        candidates = exact_index.get(_norm(latin_name), set())
        if not candidates:
            candidates = binomial_index.get(_binomial(latin_name), set())

        if not candidates:
            unmatched += 1
            continue
        if len(candidates) == 1:
            matched_unique += 1
        else:
            matched_multi += 1

        for plant_id in sorted(candidates):
            cur.execute(
                """
                UPDATE plants
                SET common_name_hungarian = ?
                WHERE id = ?
                """,
                (hungarian_name, plant_id),
            )
            if cur.rowcount:
                updated += 1

    conn.commit()

    cur.execute(
        "SELECT COUNT(*) FROM plants WHERE common_name_hungarian IS NOT NULL AND TRIM(common_name_hungarian) <> ''"
    )
    total_with_hu = cur.fetchone()[0]
    conn.close()

    print("=== Hungarian Name Import Complete ===")
    print(f"Matched (unique): {matched_unique}")
    print(f"Matched (multiple candidates): {matched_multi}")
    print(f"Unmatched: {unmatched}")
    print(f"Plant rows updated: {updated}")
    print(f"Plants with Hungarian common name now: {total_with_hu}")


if __name__ == "__main__":
    main()
