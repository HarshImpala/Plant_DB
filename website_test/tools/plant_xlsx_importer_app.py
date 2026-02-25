"""
Desktop/CLI importer for plant-page data from XLSX.

Usage:
  - GUI mode:
      python tools/plant_xlsx_importer_app.py
  - Generate template:
      python tools/plant_xlsx_importer_app.py --template data/plant_import_template.xlsx
  - Import and rebuild:
      python tools/plant_xlsx_importer_app.py --import-file path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter.scrolledtext import ScrolledText
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    ScrolledText = None


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "plants.db"
COLLECTIONS_PATH = DATA_DIR / "collections.json"
BUILD_SCRIPT = BASE_DIR / "generator" / "build_site.py"


PLANT_COLUMNS = [
    "input_name",
    "scientific_name",
    "canonical_name",
    "common_name",
    "common_name_hungarian",
    "family",
    "genus",
    "wfo_id",
    "wfo_url",
    "gbif_usage_key",
    "gbif_url",
    "wikipedia_url_english",
    "wikipedia_url_hungarian",
    "description_english",
    "description_hungarian",
    "description_hungarian_is_translated",
    "native_countries",
    "native_regions",
    "native_confidence",
    "toxicity_info",
    "garden_location",
    "image_filename",
    "image_source",
    "curator_comments",
]

# Extra relationship columns supported by this importer.
EXTRA_COLUMNS = [
    "synonyms",
    "common_names_en",
    "common_names_hu",
    "collection_slug",
]

ALL_TEMPLATE_COLUMNS = PLANT_COLUMNS + EXTRA_COLUMNS


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_bool_int(value: object) -> int:
    if value is None:
        return 0
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y"}:
        return 1
    return 0


def parse_pipe_list(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    result = []
    seen = set()
    for item in text.split("|"):
        entry = item.strip()
        if not entry:
            continue
        key = entry.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(entry)
    return result


@dataclass
class ImportStats:
    rows_seen: int = 0
    plants_upserted: int = 0
    synonyms_written: int = 0
    common_names_written: int = 0
    collections_updated: int = 0


def create_template_xlsx(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "plants"
    ws.append(ALL_TEMPLATE_COLUMNS)
    ws.append(
        [
            "Hedychium coronarium J.Koenig",
            "Hedychium coronarium J.Koenig",
            "Hedychium coronarium",
            "Butterfly-ginger",
            "feher gyomberliliom",
            "Zingiberaceae",
            "Hedychium",
            "wfo-0000435787",
            "https://www.worldfloraonline.org/taxon/wfo-0000435787",
            "7883492",
            "https://www.gbif.org/species/7883492",
            "https://en.wikipedia.org/wiki/Hedychium_coronarium",
            "",
            "Perennial flowering ginger plant...",
            "",
            0,
            "India | Nepal",
            "India | Nepal",
            "high",
            "Non-toxic to dogs, cats. (Source: ASPCA)",
            "Tropical greenhouse",
            "hedychium-coronarium.jpg",
            "manual",
            "Added via desktop importer.",
            "Amomum coronarium | Gandasulium coronarium",
            "Butterfly-ginger | White ginger lily",
            "feher gyomberliliom",
            "tropical-crop-plants",
        ]
    )

    meta = wb.create_sheet("field_notes")
    meta.append(["column", "description"])
    notes = {
        "input_name": "Required. Unique row key for upsert.",
        "canonical_name": "Used in URLs/search and collection membership.",
        "description_hungarian_is_translated": "1 or 0.",
        "synonyms": "Pipe-separated list.",
        "common_names_en": "Pipe-separated list of English common names.",
        "common_names_hu": "Pipe-separated list of Hungarian common names.",
        "collection_slug": "Optional. Adds canonical_name to data/collections.json.",
    }
    for col in ALL_TEMPLATE_COLUMNS:
        meta.append([col, notes.get(col, "Optional text field")])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def ensure_category(cursor: sqlite3.Cursor, name: str, category_type: str) -> int:
    cursor.execute(
        """
        INSERT OR IGNORE INTO categories (name, category_type)
        VALUES (?, ?)
        """,
        (name, category_type),
    )
    cursor.execute(
        "SELECT id FROM categories WHERE name = ? AND category_type = ?",
        (name, category_type),
    )
    return int(cursor.fetchone()[0])


def upsert_collection_membership(collection_slug: str, canonical_name: str) -> bool:
    if not collection_slug or not canonical_name:
        return False
    if not COLLECTIONS_PATH.exists():
        return False
    collections = json.loads(COLLECTIONS_PATH.read_text(encoding="utf-8"))
    changed = False
    for col in collections:
        if col.get("slug") != collection_slug:
            continue
        plants = col.setdefault("plants", [])
        existing = {str(x).strip().lower() for x in plants}
        if canonical_name.strip().lower() not in existing:
            plants.append(canonical_name.strip())
            changed = True
        break
    if changed:
        COLLECTIONS_PATH.write_text(
            json.dumps(collections, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    return changed


def import_xlsx(path: Path) -> ImportStats:
    if not path.exists():
        raise FileNotFoundError(f"XLSX file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    if "plants" in wb.sheetnames:
        ws = wb["plants"]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError("Excel sheet is empty.")
    columns = [str(h).strip() if h is not None else "" for h in header]
    col_index = {name: idx for idx, name in enumerate(columns)}
    if "input_name" not in col_index:
        raise ValueError("Missing required 'input_name' column.")

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    stats = ImportStats()

    for row in rows:
        stats.rows_seen += 1
        if row is None:
            continue
        input_name = normalize_text(row[col_index["input_name"]] if col_index["input_name"] < len(row) else None)
        if not input_name:
            continue

        data = {}
        for col in PLANT_COLUMNS:
            if col not in col_index:
                continue
            raw = row[col_index[col]] if col_index[col] < len(row) else None
            if col == "description_hungarian_is_translated":
                data[col] = parse_bool_int(raw)
            elif col == "gbif_usage_key":
                text = normalize_text(raw)
                data[col] = text
                if text and "gbif_url" not in col_index:
                    data["gbif_url"] = f"https://www.gbif.org/species/{text}"
            else:
                data[col] = normalize_text(raw)

        data["input_name"] = input_name
        cols_for_upsert = [c for c in PLANT_COLUMNS if c in data]
        placeholders = ", ".join("?" for _ in cols_for_upsert)
        update_clause = ", ".join(f"{c}=excluded.{c}" for c in cols_for_upsert if c != "input_name")
        cur.execute(
            f"""
            INSERT INTO plants ({", ".join(cols_for_upsert)})
            VALUES ({placeholders})
            ON CONFLICT(input_name) DO UPDATE SET
              {update_clause},
              updated_at=CURRENT_TIMESTAMP
            """,
            [data[c] for c in cols_for_upsert],
        )
        stats.plants_upserted += 1

        cur.execute("SELECT id, canonical_name FROM plants WHERE input_name = ?", (input_name,))
        plant_id, canonical_name = cur.fetchone()

        synonyms = parse_pipe_list(row[col_index["synonyms"]] if "synonyms" in col_index and col_index["synonyms"] < len(row) else None)
        if synonyms:
            cur.execute("DELETE FROM plant_synonyms WHERE plant_id = ? AND source = 'manual'", (plant_id,))
            for syn in synonyms:
                cur.execute(
                    "INSERT OR IGNORE INTO plant_synonyms (plant_id, synonym_name, source) VALUES (?, ?, 'manual')",
                    (plant_id, syn),
                )
                stats.synonyms_written += 1

        names_en = parse_pipe_list(row[col_index["common_names_en"]] if "common_names_en" in col_index and col_index["common_names_en"] < len(row) else None)
        names_hu = parse_pipe_list(row[col_index["common_names_hu"]] if "common_names_hu" in col_index and col_index["common_names_hu"] < len(row) else None)
        if names_en or names_hu:
            cur.execute("DELETE FROM plant_common_names WHERE plant_id = ? AND language IN ('en', 'hu')", (plant_id,))
            for name in names_en:
                cur.execute(
                    "INSERT OR IGNORE INTO plant_common_names (plant_id, common_name, language) VALUES (?, ?, 'en')",
                    (plant_id, name),
                )
                stats.common_names_written += 1
            for name in names_hu:
                cur.execute(
                    "INSERT OR IGNORE INTO plant_common_names (plant_id, common_name, language) VALUES (?, ?, 'hu')",
                    (plant_id, name),
                )
                stats.common_names_written += 1

        family = data.get("family")
        genus = data.get("genus")
        if family:
            category_id = ensure_category(cur, family, "family")
            cur.execute(
                "INSERT OR IGNORE INTO plant_categories (plant_id, category_id) VALUES (?, ?)",
                (plant_id, category_id),
            )
        if genus:
            category_id = ensure_category(cur, genus, "genus")
            cur.execute(
                "INSERT OR IGNORE INTO plant_categories (plant_id, category_id) VALUES (?, ?)",
                (plant_id, category_id),
            )

        collection_slug = normalize_text(
            row[col_index["collection_slug"]]
            if "collection_slug" in col_index and col_index["collection_slug"] < len(row)
            else None
        )
        if collection_slug and canonical_name and upsert_collection_membership(collection_slug, canonical_name):
            stats.collections_updated += 1

    conn.commit()
    conn.close()
    return stats


def run_build() -> None:
    subprocess.run([sys.executable, str(BUILD_SCRIPT)], cwd=str(BASE_DIR), check=True)


def format_stats(stats: ImportStats) -> str:
    return (
        f"Rows seen: {stats.rows_seen}\n"
        f"Plants upserted: {stats.plants_upserted}\n"
        f"Manual synonyms written: {stats.synonyms_written}\n"
        f"Common names written: {stats.common_names_written}\n"
        f"Collection files updated: {stats.collections_updated}"
    )


class ImporterApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        root.title("Plant XLSX Importer")
        root.geometry("860x560")

        toolbar = tk.Frame(root)
        toolbar.pack(fill=tk.X, padx=8, pady=8)

        tk.Button(toolbar, text="Generate Template", command=self.generate_template).pack(side=tk.LEFT, padx=4)
        tk.Button(toolbar, text="Import XLSX + Rebuild", command=self.import_and_build).pack(side=tk.LEFT, padx=4)

        self.log = ScrolledText(root, wrap=tk.WORD, height=30)
        self.log.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))
        self.write("Ready.\n")
        self.write(f"Project: {BASE_DIR}\n")

    def write(self, text: str) -> None:
        self.log.insert(tk.END, text)
        self.log.see(tk.END)
        self.root.update_idletasks()

    def generate_template(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save Template XLSX",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="plant_import_template.xlsx",
        )
        if not path:
            return
        out = Path(path)
        create_template_xlsx(out)
        self.write(f"Template written: {out}\n")
        messagebox.showinfo("Template created", f"Created:\n{out}")

    def import_and_build(self) -> None:
        path = filedialog.askopenfilename(
            title="Choose plant XLSX",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        src = Path(path)
        self.write(f"\nImporting: {src}\n")
        try:
            stats = import_xlsx(src)
            self.write(format_stats(stats) + "\n")
            self.write("Rebuilding site...\n")
            run_build()
            self.write("Build complete.\n")
            messagebox.showinfo("Done", "Import and rebuild finished.")
        except Exception as exc:
            self.write(f"ERROR: {exc}\n")
            messagebox.showerror("Error", str(exc))


def main() -> int:
    parser = argparse.ArgumentParser(description="Plant XLSX importer desktop tool")
    parser.add_argument("--template", type=Path, help="Write template XLSX to this path and exit")
    parser.add_argument("--import-file", type=Path, help="Import XLSX and rebuild site")
    args = parser.parse_args()

    if args.template:
        create_template_xlsx(args.template)
        print(f"Template written: {args.template}")
        return 0

    if args.import_file:
        stats = import_xlsx(args.import_file)
        print(format_stats(stats))
        run_build()
        print("Build complete.")
        return 0

    if tk is None:
        print("Tkinter is not available. Use --template or --import-file mode.")
        return 1

    root = tk.Tk()
    ImporterApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

